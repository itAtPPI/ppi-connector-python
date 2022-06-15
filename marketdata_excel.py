# Codigo propio de ppi, siempre obligatorio
from ppi_client.api.constants import ACCOUNTDATA_TYPE_ACCOUNT_NOTIFICATION, ACCOUNTDATA_TYPE_PUSH_NOTIFICATION, \
    ACCOUNTDATA_TYPE_ORDER_NOTIFICATION
from ppi_client.models.account_movements import AccountMovements
from ppi_client.models.bank_account_request import BankAccountRequest
from ppi_client.models.foreign_bank_account_request import ForeignBankAccountRequest, ForeignBankAccountRequestDTO
from ppi_client.models.cancel_bank_account_request import CancelBankAccountRequest
from ppi_client.models.order import Order
from ppi_client.ppi import PPI
from ppi_client.models.order_budget import OrderBudget
from ppi_client.models.order_confirm import OrderConfirm
from ppi_client.models.disclaimer import Disclaimer
from ppi_client.models.investing_profile import InvestingProfile
from ppi_client.models.investing_profile_answer import InvestingProfileAnswer
from ppi_client.models.instrument import Instrument
from datetime import datetime, timedelta
import asyncio
import json
import traceback
import os
from openpyxl import Workbook

# Change sandbox variable to False to connect to production environment
ppi = PPI(sandbox=False)


def main():
    try:
        # Change login credential to connect to the API
        ppi.account.login('<user key>', '<user secret>')

        # Obtengo las cotizaciones historicas
        print("\nSearching MarketData")
        market_data = ppi.marketdata.search("GGAL", "Acciones", "A-48HS", datetime(2021, 1, 1), datetime(2021, 12, 31))

        # Creo el excel
        workbook = Workbook()
        sheet = workbook.active

        # Completo los encabezados
        sheet.cell(row=1, column=1).value = "Fecha"
        sheet.cell(row=1, column=2).value = "Cotizacion"
        sheet.cell(row=1, column=3).value = "Volumen"
        sheet.cell(row=1, column=4).value = "Apertura"
        sheet.cell(row=1, column=5).value = "Minimo"
        sheet.cell(row=1, column=6).value = "Maximo"

        r = 1
        # LLeno el excel
        for ins in market_data:
            r = r + 1
            sheet.cell(row=r, column=1).value = ins['date']
            sheet.cell(row=r, column=2).value = ins['price']
            sheet.cell(row=r, column=3).value = ins['volume']
            sheet.cell(row=r, column=4).value = ins['openingPrice']
            sheet.cell(row=r, column=5).value = ins['min']
            sheet.cell(row=r, column=6).value = ins['max']

        #Guardo el excel
        workbook.save(filename="marketdata.xlsx")
        print("\nExportacion realizada con exito")

    except Exception as message:
        print(datetime.now())
        print(message)

    input('\nPress Enter to exit')

if __name__ == '__main__':
    main()
